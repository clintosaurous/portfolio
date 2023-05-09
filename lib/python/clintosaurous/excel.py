#!/opt/clintosaurous/venv/bin/python3 -Bu

"""
Module to work with Excel spreadsheets.
"""


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re


VERSION = "1.1.0"
LAST_UPDATE = "2022-09-28"


def _column_width(string, bold=False):

    """
    Attempt to calculate the width in Excel sizing of the given string.
    """

    if not isinstance(string, str):
        string = str(string)

    # padding width for the right and left side of text in a cell
    cell_side_padding = 0.6

    # average character width in Excel sizing
    if bold:
        # wide character width in Excel sizing is 1.34
        #   difference between average and wide is 0.16
        wide_char_diff = 0.2
        # short character width in Excel sizing is 0.75
        #   difference between average and wide is 0.43
        short_char_diff = 0.5
        avg_char_len = 1.4
    else:
        # wide character width in Excel sizing is 1.34
        #   difference between average and wide is 0.16
        wide_char_diff = 0.1
        # short character width in Excel sizing is 0.75
        #   difference between average and wide is 0.43
        short_char_diff = 0.45
        avg_char_len = 1.15

    # determine the count of wide characters in the string
    wide_char_cnt = len(re.findall(r'[A-CE-HK-Zmw_\@\$\%\&]', string))
    if not wide_char_cnt:
        wide_char_cnt = 0

    # determine the count of short characters in the string
    short_char_cnt = \
        len(re.findall(r'[fiIjJlrst!\(\)\-\[\]\{\}\|:;,\.\']', string))
    if not short_char_cnt:
        short_char_cnt = 0

    # width if all characters were average width in string
    width = len(string) * avg_char_len + cell_side_padding
    # add width for all characters larger than average width in string
    width = width + wide_char_cnt * wide_char_diff
    # subtract width for all characters shorter than average width in string
    width = width - short_char_cnt * short_char_diff

    # set a maximum width of 75
    if width > 75:
        width = 75

    return width

# End: _column_width()


def generate(file_name, sheets):

    """
    excel.generate(file_name, sheets)

    Creates an Excel workbook to the given file, with optional headers, and
    the supplied rows.

    sheets: List of each sheet to be added to the workbook. Each row contains
    the sheet specific information. Sheets are added in the order they are in
    the list.

    Each sheet is a dictionary with the below keys:

        title:      The sheet name on the tab in the workbook.
        headings:   A list of column headings.
        rows:       A list of list rows of data.
    """

    wb = Workbook()
    bg_heading = PatternFill(
        start_color="DDDDDD",
        end_color="DDDDDD",
        fill_type="solid"
    )
    bg_even_row = PatternFill(
        start_color="F3F9FB",
        end_color="F3F9FB",
        fill_type="solid"
    )
    bg_odd_row = PatternFill(
        start_color="D4E2F4",
        end_color="D4E2F4",
        fill_type="solid"
    )
    font_bold = Font(bold=True)

    for sheet_num in range(len(sheets)):
        sheet = sheets[sheet_num]
        col_widths = []
        col_types = {}

        try:
            title = sheet["title"]
        except KeyError:
            title = f'Sheet{sheet_num + 1}'

        if sheet_num:
            ws = wb.create_sheet(title)
        else:
            ws = wb.active
            ws.title = title

        try:
            headings = True
            ws.append(sheet["headings"])
            ws.freeze_panes = 'A2'
            col_widths = [_column_width(h, True) for h in sheet["headings"]]
            col_types = [None for i in range(len(sheet["headings"]))]

        except KeyError:
            headings = False
            col_widths = []
            col_types = []

        for row in sheet["rows"]:
            for col_num in range(len(row)):
                row[col_num] = row[col_num].strip()
                col_width = _column_width(row[col_num], False)
                try:
                    if col_width > col_widths[col_num]:
                        col_widths[col_num] = col_width
                except IndexError:
                    col_widths.append(col_width)

                if (
                    isinstance(row[col_num], int)
                    or isinstance(row[col_num], float)
                    or (
                        isinstance(row[col_num], str)
                        and (
                            re.match(r'^[\d,]+$', row[col_num])
                            or re.match(r'^[\d,]+\.\d+$', row[col_num])
                            or re.match(r'^[\d.]+\%$', row[col_num])
                            or re.match(r'^[\d,]+\.\d+\%$', row[col_num])
                        )
                    )
                ):
                    try:
                        col_types[col_num]
                    except IndexError:
                        col_types.append('number')

                else:
                    try:
                        col_types[col_num] = 'string'
                    except IndexError:
                        col_types.append('string')

        if headings:
            row_offset = 2
            for head_num in range(len(sheet["headings"])):
                cell = get_column_letter(head_num + 1) + "1"
                ws[cell].fill = bg_heading
                ws[cell].font = font_bold
        else:
            row_offset = 1

        for row_num in range(len(sheet["rows"])):
            row = sheet["rows"][row_num]
            ws.append(row)

            for col_num in range(len(row)):
                cell = \
                    get_column_letter(col_num + 1) + str(row_num + row_offset)

                col_len = _column_width(row[col_num], False)
                try:
                    if col_len > col_widths[col_num]:
                        col_widths[col_num] = col_len
                except IndexError:
                    col_widths.append(col_len)

                if col_types[col_num] == 'number':
                    align = ws[cell].alignment.copy(horizontal='right')
                    ws[cell].alignment = align
                else:
                    align = ws[cell].alignment.copy(horizontal='left')
                    ws[cell].alignment = align

        for col_num in range(len(col_widths)):
            ws.column_dimensions[get_column_letter(col_num + 1)].width = \
                col_widths[col_num]

        if headings:
            ws.auto_filter.ref = ws.dimensions

    wb.save(file_name)

# End: generate()
